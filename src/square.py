#!/usr/bin/env python3
# Generate movement and scores for "Square Mitchell"
#   A PDF with Roster and Score sheets
#   An Excel spreadsheet to enter the results and calculate the scores
# A Square Mitchell is designed for 8 pairs playing Mitchell without using a relay table.
# NS, as normal, stay stationary.
# EW, however, sometimes move diagonal.
# The only option is the number of boards to play for each round
#
# Reuse code for general Mitchell movement
#
import argparse
import logging
from maininit import setlog
from openpyxl import load_workbook
from mitchell import Mitchell

class Square(Mitchell):
    def __init__(self, log, b, f):
        super().__init__(log, 8, b, f)

    def go(self):
        # the sequence of calls is important
        self.loadSetupSheet()
        self.roster()
        self.roundTab() # build data structure for late
        self.boardTab()
        self.results()
        self.ScoreTable()
        self.idTags()
        self.Pickups()  # PDF only
        self.Tables(True)  # PDF only
        self.Travelers()  # PDF only
        self.Journal()  # PDF only
        self.save()
        return

    # Also setup "boardData" for subsequent calls
    def roundTab(self):
        self.log.debug('Saving by Round')
        headers = ['Round', 'Table', 'NS', 'EW', 'Board', 'Vul', 'Contract', 'By', 'Result', 'NS', 'EW']
        sh, row = self.contractHeaders(headers, 'By Round', ['Scores'])
        roundData = {}
        self.boardData = {}
        for t,tData in self.sqSetup.items():
            for r in tData:
                if r['Round'] not in roundData:
                    roundData[r['Round']] = []
                roundData[r['Round']].append((t, r['NS'], r['EW'], r['Board']))
        for r in sorted(roundData.keys()):
            sh.cell(row, 1).value = r + 1
            for t in sorted(roundData[r], key=lambda x: x[0]):
                # Round, NS, EW
                sh.cell(row, 2).value = t[0] + 1
                sh.cell(row, 3).value = t[1]
                sh.cell(row, 4).value = t[2]
                for j in range(1,5):
                    sh.cell(row, j).alignment = self.centerAlign
                # A row for each board
                for i in t[3]:
                    sh.cell(row, 5).value = i+1
                    sh.cell(row, 6).value = self.vulLookup(i)
                    sh.cell(row, 5).alignment = sh.cell(row, 6).alignment = self.centerAlign
                    # Caceh boardData for future uses
                    if i not in self.boardData:
                        self.boardData[i] = []
                    self.boardData[i].append((r, t[0], t[1], t[2]))

                    # fake scores for debugging
                    if self.fake:
                        self.fakeScore(sh, row, 10)
                    row += 1
                # draw lines for visibility
                for c in range(len(headers) - 1):
                    sh.cell(row-1, c+2).border = self.bottomLine
            for c in range(len(headers)):
                sh.cell(row-1, c+1).border = self.bottomLine


    def loadSetupSheet(self):
        self.sqSetup = {
            # Primary key is the table number
            # Pair numbering in this data is separated by NS/EW
            # Each "board set" is n boards, as dedicated by command line argument
            0: [{'Round': 0, 'NS': 1, 'EW': 1, 'Board': 0},   # round, NS, EW, boardSet #
                {'Round': 1, 'NS': 1, 'EW': 2, 'Board': 3},
                {'Round': 2, 'NS': 1, 'EW': 3, 'Board': 2},
                {'Round': 3, 'NS': 1, 'EW': 4, 'Board': 1},],
            1: [{'Round': 0, 'NS': 2, 'EW': 2, 'Board': 1},
                {'Round': 1, 'NS': 2, 'EW': 1, 'Board': 2},
                {'Round': 2, 'NS': 2, 'EW': 3, 'Board': 3},
                {'Round': 3, 'NS': 2, 'EW': 4, 'Board': 0},],
            2: [{'Round': 0, 'NS': 3, 'EW': 3, 'Board': 2},
                {'Round': 1, 'NS': 3, 'EW': 4, 'Board': 1},
                {'Round': 2, 'NS': 3, 'EW': 2, 'Board': 0},
                {'Round': 3, 'NS': 3, 'EW': 1, 'Board': 3},],
            3: [{'Round': 0, 'NS': 4, 'EW': 4, 'Board': 3},
                {'Round': 1, 'NS': 4, 'EW': 3, 'Board': 0},
                {'Round': 2, 'NS': 4, 'EW': 1, 'Board': 1},
                {'Round': 3, 'NS': 4, 'EW': 2, 'Board': 2}]}
        for tbl in self.sqSetup.values():
            for r in tbl:
                r['NS'] = (r['NS'] - 1) * 2 + 1
                r['EW'] *= 2
                r['Board'] = [r['Board']*self.boards + x for x in range(self.boards)]
        self.tables = len(self.sqSetup.keys())  # used by Mitchell

    def idTags(self):
        idData = {}
        for t, d in self.sqSetup.items():
            for r in d:
                if r['NS'] not in idData:
                    idData[r['NS']] = []
                if r['EW'] not in idData:
                    idData[r['EW']] = []
                idData[r['NS']].append((r['Round'], t, r['NS'], r['EW']))
                idData[r['EW']].append((r['Round'], t, r['NS'], r['EW']))
        self.idTagsByData(idData)

    # Data {pair #: [(round, table, NS, EW), ...], ...}
    # A page holds 4 sets of "id tags", one for each person of each pair
    def idTagsByData(self, data):
        nTagsPage = len(data) if len(data) <= 4 else 4
        cHeight = self.pdf.eph / nTagsPage
        cWidth = self.pdf.w / 2
        leftMargin = self.pdf.margin * 2
        tags = 0
        colW = []
        hdrs = ['Round', 'Table', 'NS', 'EW']
        self.pdf.setHeaders(leftMargin, hdrs, colW)
        for id in sorted(data.keys()):
            if tags % nTagsPage == 0:
                self.pdf.add_page() # no header/footer
                y = self.pdf.margin
            rData = sorted(data[id], key=lambda x: x[0])
            for half in range(2):   # two identical tags for each person of the pair
                self.pdf.set_font(self.pdf.serifFont, size=self.pdf.headerPt)
                self.pdf.set_xy(leftMargin+cWidth*half, y)
                self.pdf.cell(text=f"Pair: {self.pairID(id)}")
                h = self.pdf.lineHeight(self.pdf.font_size_pt)
                ty = y + h
                self.pdf.set_xy(leftMargin+cWidth*half, ty)
                self.pdf.set_font(self.pdf.sansSerifFont, style='B', size=self.pdf.smallPt)
                h = self.pdf.lineHeight(self.pdf.font_size_pt)
                for i in range(len(hdrs)):
                    self.pdf.cell(colW[i], h, text=hdrs[i], align='C', border=1)
                ty +=  h
                for r in rData:
                    self.pdf.set_xy(leftMargin+cWidth*half, ty)
                    for i in range(len(hdrs)):
                        if i <= hdrs.index('Table'):
                            txt = f"{r[i]+1}"
                        else:
                            txt = f"{self.pairN(r[i])}"
                        self.pdf.cell(colW[i], h, text=txt, align='C', border=1)
                    ty += h
            y += cHeight
            tags += 1
        return
           
    # Output into filesystem
    def save(self):
        import os
        here = os.path.dirname(os.path.abspath(__file__))
        fn = f'{here}/../squarex{self.boards}'
        self.wb.save(f'{fn}.xlsx')
        self.pdf.output(f'{fn}.pdf')
        print(f'Saved {fn}.{{xlsx,pdf}}')

if __name__ == '__main__':
    log = setlog('mitchell', None)
    def even_type(value):
        ivalue = int(value)
        if ivalue % 2 != 0:
            raise argparse.ArgumentTypeError(f"{value} is not an even number")
        return ivalue

    parser = argparse.ArgumentParser()
    parser.add_argument('-d', '--debug', type=str, default='INFO', help='Debug level, INFO, DEBUG, ERROR')
    parser.add_argument('-b', '--boards', type=int, choices=range(2,7), default=3, help='Boards per round')
    parser.add_argument('-f', '--fake', type=bool, default=False, help='Fake scores to test the spreadsheet')
    args = parser.parse_args()
    for l in [['INFO', logging.INFO], ['DEBUG', logging.DEBUG], ['ERROR', logging.ERROR]]:
        if args.debug.upper() == l[0]:
            log.setLevel(l[1])
            break
    sq = Square(log, args.boards, args.fake)
    # A match has n rounds, each round has m boards, divided into two halves, each half of the boards
    sq.go()