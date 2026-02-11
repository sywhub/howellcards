#!/usr/bin/env python3
# Generate Howell movement placards and scoring spreadsheet based on the number of pairs
# The tournament setups were pre-generated and stored as JSON.
# The algorithm of generating those data is a separate program, as those generations take time.

# --pair #: generate sheets for the specific pair #, if absent do all of them
# --fake: fake results to test the scoring mechanism in the spreadsheet
# --debug <DEBUG LEVEL>: used only by the developer

# to do: do Google sheet instead of Microsoft Excel
#        Smooother board transitions

import argparse
import pdf
import os
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import logging
import jsonIO
from maininit import setlog
from docset import PairGames

class Howell(PairGames):
    def __init__(self, log, toFake, pairs, tourney):
        super().__init__(log)
        self.fake = toFake
        self.pdf = pdf.PDF()
        self.wb = Workbook()
        self.pairs = pairs
        self.tourneyData = tourney
        self.init()
        return

    def save(self):
        here = os.path.dirname(os.path.abspath(__file__))
        fn = f'{here}/../howell{self.pairs}{"xF" if self.fake else ""}'
        self.wb.save(f'{fn}.xlsx')
        self.pdf.output(f'{fn}.pdf')
        print(f'Saved {fn}.{{xlsx,pdf}}')

    def pairN(self, n):
        return n if n != 0 else self.SITOUT

    def pairID(self, n):
        return f"{n if n != 0 else self.SITOUT}"
	
    def ifSitout(self, t, ns, ew):
        return ns == 0

    # string to enumerate a "board set" into individual decks
    def boardSet(self, bIdx):
        bds = self.boardList(bIdx)
        str = ''
        for i in bds:
            str += f'{i+1}'
            if i != bds[-1]:
                str += ' & '
        return str

    # probably should be part of the constructor
    # Initialize some state.
    # Create meta and roster sheets
    def init(self):
        if self.pairs <= 6:
            self.decks = 3
        else:
            self.decks = 2
        self.boardData = {}
        for r in range(len(self.tourneyData['Arrangement'])):
            for t in range(len(self.tourneyData['Arrangement'][r])):
                tbl = self.tourneyData['Arrangement'][r][t]
                for b in self.boardList(tbl['Board']):
                    if b not in self.boardData:
                        self.boardData[b] = []
                    self.boardData[b].append([r, t, tbl['NS'], tbl['EW']])
        self.initRounds()
        self.checkBoardData()
        self.tables = len(self.roundData[0])
        nRound = self.tourneyData['Rounds']

        # meta data
        self.metaData = {'Title': 'Howell Tournament',
            'Info': [['Pairs', self.pairs], ['Tables',int((self.pairs + (self.pairs % 2))/ 2)],
            ['Rounds',nRound], ['Boards per round',self.decks], ['Total Boards to play', self.decks*nRound]]}

        self.pdf.HeaderFooterText(f'{self.notice} {datetime.date.today().strftime("%b %d, %Y")}.',
           f'Howell Movement for {self.pairs} Pairs')

    # Present the same data table-oriented
    def movementTables(self):
        self.log.debug('Saving by Table')
        rounds = self.tourneyData['Arrangement']
        nTbl = len(rounds[0])
        nRounds = len(rounds)
        # tbl#: {'nRound': # of rounds, r: ({'NS': ns, 'EW': ew, 'Board': board set #}, "boards set")}
        moveData = {}
        # iterate by table then by round
        for tbl in range(nTbl):
            moveData[tbl] = {}
            for r in range(nRounds):
                if r != nRounds - 1:
                    for side in ['NS', 'EW']:
                        # build a reverse lookup of "side: table" of next round
                        next = {v[side]: k for k,v in enumerate(rounds[r+1])}
                        # look up the pair im that side's lookup
                        if rounds[r][tbl]['NS'] in next.keys():
                            moveData[tbl]['nsNext'] = (next[rounds[r][tbl]['NS']], side)
                        if rounds[r][tbl]['EW'] in next.keys():
                            moveData[tbl]['ewNext'] = (next[rounds[r][tbl]['EW']], side)
        nsText = []
        ewText = []
        for t in sorted(moveData.keys()):    # tables
            if t == 0:
                nsText.append('Stay Stationary')
            else:
                nsText.append(f'Move To Table {moveData[t]['nsNext'][0]+1} {moveData[t]['nsNext'][1]}')
            ewText.append(f'Move To Table {moveData[t]['ewNext'][0]+1} {moveData[t]['ewNext'][1]}')
        self.Tables(nsText, ewText)

    # Roster sheet
    # Also the final result
    def rosterSheet(self):
        self.log.debug('Creating Roster Sheet')

        sh = self.wb.active
        sh.title = 'Roster'
        row = self.sheetMeta(sh, self.metaData) + 2
        sh.cell(row, 1).value =  'Pairs'
        sh.cell(row, 1).font = self.HeaderFont
        sh.cell(row, 1).alignment = self.centerAlign
        sh.merge_cells(f'{sh.cell(row,1).coordinate}:{sh.cell(row,3).coordinate}')
        sh.cell(row, 4).value = 'MP'
        sh.cell(row, 4).font = self.HeaderFont
        sh.cell(row, 4).alignment = self.centerAlign
        sh.cell(row, 5).value = 'IMP'
        sh.cell(row, 5).font = self.HeaderFont
        sh.cell(row, 5).alignment = self.centerAlign
        row += 1

        totalPlayed = int((self.pairs + self.pairs % 2) / 2) * self.decks * (self.pairs - 1)
        for i in range(self.pairs):
            sh.cell(i+row, 1).value = i+1
            sh.cell(i+row, 1).font = self.HeaderFont
            sh.cell(i+row, 1).alignment = self.centerAlign
            sh.cell(i+row, 2).value = self.placeHolderName()
            sh.cell(i+row, 3).value = self.placeHolderName()
            sh.column_dimensions['B'].width = 25
            sh.column_dimensions['C'].width = 25
            IMPsum1 = f"=SUMIF('By Board'!$D$3:$D${totalPlayed+2},\"={i+1}\",'By Board'!$L$3:$L${totalPlayed+2})"
            MPsum1 = f"=SUMIF('By Board'!$D$3:$D${totalPlayed+2},\"={i+1}\",'By Board'!$N$3:$N${totalPlayed+2})"
            if self.pairs % 2 != 0 or i != self.pairs - 1:
                IMPsum2 = f"SUMIF('By Board'!$E$3:$E${totalPlayed+2},\"={i+1}\",'By Board'!$M$3:$M${totalPlayed+2})"
                MPsum2 = f"SUMIF('By Board'!$E$3:$E${totalPlayed+2},\"={i+1}\",'By Board'!$O$3:$O${totalPlayed+2})"
            else:
                IMPsum2=0
                MPsum2=0
            sh.cell(i+row, 4).value = f"{MPsum1}/{self.decks*(self.pairs-1)}+{MPsum2}/{self.decks*(self.pairs-1)}"
            sh.cell(i+row, 4).number_format = '0.0%'
            sh.cell(i+row, 5).value = f"{IMPsum1}+{IMPsum2}"
            sh.cell(i+row, 5).number_format = '#0.00'
        
        IMPRow = self.pairs + row + 2
        sh.cell(IMPRow, 1).value = 'To sort, remove single quote below'
        IMPRow += 1
        sh.cell(IMPRow,1).value = 'IMP Ranking'
        sh.cell(IMPRow,1).alignment = self.centerAlign
        sh.cell(IMPRow,1).font = self.HeaderFont
        sh.merge_cells(f'{self.rc2a1(IMPRow, 1)}:{self.rc2a1(IMPRow,5)}')
        IMPRow += 1
        sh.cell(IMPRow,1).value = f"'=SORT({self.rc2a1(row, 1)}:{self.rc2a1(row+self.pairs-1,5)},4,-1)"
        for i in range(self.pairs):
            sh.cell(IMPRow+i,4).number_format = "#0.00"
            sh.cell(IMPRow+i,5).number_format = "0.00%"
        MPRow = self.pairs + IMPRow + 4
        sh.cell(MPRow,1).value = 'MP Ranking'
        sh.cell(MPRow,1).alignment = self.centerAlign
        sh.cell(MPRow,1).font = self.HeaderFont
        sh.merge_cells(f'{self.rc2a1(MPRow, 1)}:{self.rc2a1(MPRow,5)}')
        MPRow += 1
        sh.cell(MPRow,1).value = f"'=SORT({self.rc2a1(row, 1)}:{self.rc2a1(row+self.pairs-1,5)},5,-1)"
        for i in range(self.pairs):
            sh.cell(MPRow+i,4).number_format = "#0.00"
            sh.cell(MPRow+i,5).number_format = "0.00%"

        # Check to make sure IMPs add up to zero
        ft = Font(bold=True,color="FF0000")
        topBorder = Border(top=Side(style='thin', color="FF0000"))
        sh.cell(self.pairs+row, 4).value=f'=AVERAGE(D{row}:D{row+self.pairs-1})'
        sh.cell(self.pairs+row, 4).number_format = '0.00%'
        sh.cell(self.pairs+row, 5).value=f'=SUM(E{row}:E{row+self.pairs-1})'
        sh.cell(self.pairs+row, 5).number_format = '#0.00'
        sh.cell(self.pairs+row, 4).font = ft
        sh.cell(self.pairs+row, 4).border = topBorder
        sh.cell(self.pairs+row, 5).font = ft
        sh.cell(self.pairs+row, 5).border = topBorder

    # Sign-up sheet
    def rosterPDF(self):
        self.pdf.add_page()
        self.pdf.headerFooter()
        self.pdf.meta(self.metaData)
        self.pdf.set_font(self.pdf.serifFont, style='B', size=self.pdf.rosterPt) 
        h = self.pdf.lineHeight(self.pdf.font_size_pt)
        title = 'Player Pairs'
        x = self.pdf.setHCenter(self.pdf.get_string_width(title))
        y = self.pdf.get_y() + 2 * h
        self.pdf.set_xy(x, y)
        self.pdf.cell(text=title)
        widths = [1, 2, 2]
        y +=  h
        leftM = (self.pdf.w - sum(widths)) / 2
        self.pdf.set_xy(leftM, y)
        self.pdf.set_font(self.pdf.sansSerifFont, size=(self.pdf.bigPt if self.pairs < 19 else self.pdf.linePt)) 
        h = self.pdf.lineHeight(self.pdf.font_size_pt)
        for i in range(self.pairs):
            self.pdf.set_xy(leftM, y)
            self.pdf.cell(widths[0], h, text=f'{self.pairN(i+1)}', align='C', border=1)
            self.pdf.cell(widths[1], h, text='', align='C', border=1)
            self.pdf.cell(widths[2], h, text='', align='C', border=1)
            y += h

    def go(self):
        self.rosterSheet()
        self.boardTab()
        self.roundTab()
        self.IMPTable()
        self.ScoreTable()

        self.pdf.instructions(self.log, "instructions.txt")
        self.rosterPDF()
        self.idTags()
        self.movementTables()
        self.Travelers()
        self.Pickups()
        self.Journal()
        self.save()

def howellFromJson(log, pairs, fake, jsonfile):
    jIO = jsonIO.JsonIO(pairs, log)
    tourney = jIO.load(jsonfile)
    if tourney:
        doc = Howell(log, fake, pairs, tourney)
        doc.go()

if __name__ == '__main__':
    log = setlog('howell', None)
    parser = argparse.ArgumentParser()
    parser.add_argument('-p', '--pair', type=int, choices=range(5,15), help='# of pairs in the tournament')
    parser.add_argument('-f', '--fake', action='store_true', help='Fake scores to test the spreadsheet')
    parser.add_argument('-d', '--debug', type=str, default='INFO')
    parser.add_argument('-j', '--jsonfile', type=str)
    args = parser.parse_args()
    for l in [['INFO', logging.INFO], ['DEBUG', logging.DEBUG], ['ERROR', logging.ERROR]]:
        if args.debug.upper() == l[0]:
            log.setLevel(l[1])
            break

    if args.pair: 
        howellFromJson(log, args.pair, args.fake, args.jsonfile)
    elif args.pair is None:
        for p in range(5,15):
            howellFromJson(log, p, args.fake, args.jsonfile)

