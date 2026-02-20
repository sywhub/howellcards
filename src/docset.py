#!/usr/bin/env python3
# Base classes for bridge tournament arrangements
# These classes depend on the data set generated offline.
#
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.errors import IgnoredError
import random

# Duplicate Bridge
class DupBridge:
    def __init__(self, log):
        self.log = log
        self.HeaderFont = Font(bold=True, size=14)
        self.centerAlign = Alignment(horizontal='center')
        self.trumps = ('D/C', 'H/S', 'NT')  
        self.thinLine = Side(style='thin', color="000000")
        self.mediumLine = Side(style='medium',color="000000")
        self.bottomLine = Border(bottom=self.thinLine)
        self.thinTop = Border(top=self.thinLine)
        self.thinLeft = Border(left=self.thinLine)
        self.HeaderFont = Font(bold=True, size=14)
        self.notice = 'For public domain. No rights reserved. Generated on'
        self.fake = False

    # IMP conversion table
    def IMPTable(self):
        IMPRanges = [0, 20, 50, 90, 130, 170, 220, 270, 320, 370, 430, 500, 600, \
                    750, 900, 1100, 1300, 1500, 1750, 2000, 2250, 2500, 3000, 3500, 4000, 10010]
        sh = self.wb.create_sheet('IMP Table')
        row = self.headerRow(sh, ['From', 'To', 'IMP'])
        for i in range(0, len(IMPRanges)-1):
            sh.cell(row, 1).value = IMPRanges[i] 
            sh.cell(row, 2).value = IMPRanges[i+1] - 10
            sh.cell(row, 3).value = i
            row += 1
        return row

    # (Default) 1st row as the header
    # return the next row for data
    def headerRow(self, sh, headers, row = 1):
        col = 1
        for h in headers:
            sh.cell(row, col).font = self.HeaderFont
            sh.cell(row, col).alignment = self.centerAlign
            if type(h) is str:
                sh.cell(row, col).value = h
            else:
                sh.cell(row, col).value = h[0]
                mRange = f'{self.rc2a1(row, col)}:{self.rc2a1(row, col+h[1]-1)}'
                sh.merge_cells(mRange)
                col += h[1] - 1
            col += 1
        if 'Contract' in headers:
            contractCol = headers.index('Contract')
            sh.column_dimensions[chr(ord('A')+contractCol)].width = 30;

        return row + 1

    # A sheet for easy scoring
    def ScoreTable(self):
        sh = self.wb.create_sheet('Scoring Table')
        for c in [3, 11]:
            v = ['Not Vulnerable', 'Vulnerable']
            for x in range(2):
                sh.cell(1, c+x*3).value = v[x]
                sh.cell(1, c+x*3).font = self.HeaderFont
                sh.merge_cells(f'{sh.cell(1, c+x*3).coordinate}:{sh.cell(1, c+x*3+2).coordinate}')
                sh.cell(1, c+x*3).alignment = self.centerAlign
        headers = ['Contract', 'Made', '', 'X', 'XX', '', 'X', 'XX']
        row = self.headerRow(sh, headers, 2)
        self.scorePenalty(sh, row, len(headers)+2, headers[2:])
        for i in range(1,8):
            for trump in self.trumps:
                sh.cell(row, 1).value = f'{i} {trump}'
                sh.cell(row, 1).font = self.HeaderFont
                sh.cell(row, 1).alignment = self.centerAlign
                for j in range(0, 8-i):
                    sh.cell(row, 2).value = j+i
                    for k in range(3):
                        sh.cell(row, 3+k).value = self.score(i, trump, j, False, k)
                        sh.cell(row, 6+k).value = self.score(i, trump, j, True, k)
                        sh.cell(row, 3+k).number_format = sh.cell(row, 6+k).number_format = "#0"
                    row += 1

    # This is based on the rules for duplicate bridge
    # We use table lookup extensively
    def score(self, level, trumpSuit, res, vul, dbl):
        baseScores = [[20],[30],[40,30]]
        overDblBonus = [100, 200]
        gameBonus = [300, 500]
        slamBonus = [500, 750]
        gSlamBonus = [1000, 1500]
        dblBonus = 50
        gameThreshold = 100
        partialBonus = 50

        dblMul = 2**dbl # 2 to the power of "dbl" which is 0, 1, or 2
        vulIdx = 1 if vul else 0
        score = 0

        # Pick the right table
        tbl = baseScores[self.trumps.index(trumpSuit)]
        lTbl = len(tbl) - 1
        for c in range(level):
            if c < lTbl:
                score += tbl[c]
            else:
                score += tbl[lTbl]
        score *= dblMul
        if score < gameThreshold:
            score += partialBonus
        else:   # made game
            score += gameBonus[vulIdx]
        if level == 6:
            score += slamBonus[vulIdx]
        elif level == 7:
            score += gSlamBonus[vulIdx]
        score += dblBonus * dbl
        overTricks = res * tbl[lTbl] * dblMul
        if overTricks > 0 and dbl > 0:
            overTricks = res * overDblBonus[vulIdx] * dbl
        score += overTricks
        return score

    # The table for failing the contract
    def scorePenalty(self, sh, row, col, headers):
        penaltyTbl = [[50], [100], [100, 200, 200, 300], [200, 300]]         
        headers.insert(0, 'Down by')
        for i in range(len(headers)):
            sh.cell(row-1, col+i).value = headers[i]
            sh.cell(row-1, col+i).font = self.HeaderFont
            sh.cell(row-1, col+i).alignment = self.centerAlign

        sh.cell(row, col).value = -1
        sh.cell(row, col+1).value = -penaltyTbl[0][0]
        sh.cell(row, col+2).value = -penaltyTbl[2][0]
        sh.cell(row, col+3).value = f'={chr(ord('B')+col)}{row}*2'
        sh.cell(row, col+4).value = -penaltyTbl[1][0]
        sh.cell(row, col+5).value = -penaltyTbl[3][0]
        sh.cell(row, col+6).value = f'={chr(ord('B')+col+3)}{row}*2'
        row += 1
        for down in range(2,14):
            sh.cell(row, col).value = -down
            sh.cell(row, col+1).value = f'={sh.cell(row-1,col+1).coordinate}-{penaltyTbl[0][0]}'
            sh.cell(row, col+2).value = f'={sh.cell(row-1,col+2).coordinate}-{penaltyTbl[2][down - 1 if down <= 4 else 3]}'
            sh.cell(row, col+3).value = f'={sh.cell(row,col+2).coordinate}*2'
            sh.cell(row, col+4).value = f'={sh.cell(row-1,col+1).coordinate}-{penaltyTbl[1][0]}'
            sh.cell(row, col+5).value = f'={sh.cell(row-1,col+2).coordinate}-{penaltyTbl[3][1]}'
            sh.cell(row, col+6).value = f'={sh.cell(row,col+2).coordinate}*2'
            row += 1

    def placeHolderName(self):
        return f'Name {random.randint(11,90)}'

    # Row/Colomn to "A1" style
    # If "sheet" is avaiable, could have use the "cooridate" method.
    def rc2a1(self, r, c):
        col = ''
        c -= 1
        if c >= 26:
            col = chr(ord('A') + c // 26 - 1)
        col += chr(c%26+ord('A'))
        return f"{col}{r}"

    # Vulnerability of a board (zero-based)
    # Vulnerability is a rotation of 16 boards.  First 4 is the default.
    # Then each subsequent 4 is a left-shift rotation of the values.
    # After 16 boards, it's back to the original.
    def vulLookup(self, bidx):
        vulShift = bidx // 4
        return ['None', 'NS', 'EW', 'Both'][(bidx + vulShift) % 4]

# Howell and Mitchell tournaments
# We really arrange all tournaments to be "pair" oriented.
# The tournament is arranged by two data structures "boardData" and "roundData"
# Then they have different scoring/ranking methodologies.
class PairGames(DupBridge):
    def __init__(self, log):
        # We expect "self.tables" and "self.decks" to be done by the child class
        super().__init__(log)
        self.noChangeFont = Font(bold=True, italic=True, color='FF0000')
        self.SITOUT = "Sit-Out"
        self.roundData = {} # meant to be write-once
        self.boardData = {} # meant to be write-once

    # Placeholder functions, expect to be over-written by child classes
    # Turn internal pair number to human readable value
    def ifSitout(self, t, ns, ew):
        return False

    def pairN(self, n):
        return n + 1

    # turn pair number to string
    def pairID(self, n):
        return f"{self.pairN(n)}"
    
    # Return a list of actual board numbers (zero-based) from "board set" (zero-based)
    def boardList(self, bIdx):
        return [self.decks*bIdx+x for x in range(self.decks)]

    # Insert a score to check calculations
    # Probably could have been more sophisticated
    def fakeScore(self, sh, row, col, avgProb=0.9):
        if random.random() < avgProb:
            pickSide = col if random.random() >= 0.5 else col+1
            score = random.randint(2,80)*10
            sh.cell(row, pickSide).value = score
        else:
            sh.cell(row, col).value = 'Avg'
            sh.cell(row, col+1).value = 'Avg'

    # Construct "roundData" from "boardData".
    def initRounds(self):
        # A convenient restructure
        for b,bset in self.boardData.items():
            for s in bset:
                if s[0] not in self.roundData:   # round
                    self.roundData[s[0]] = {}
                if s[1] not in self.roundData[s[0]]: # table
                    self.roundData[s[0]][s[1]] = {'NS': s[2], 'EW': s[3], 'Board': []}
                self.roundData[s[0]][s[1]]['Board'].append(b)

    # Generate spreadsheet tab of "By Round" based on "roundData"
    def roundTab(self):
        self.log.debug('Saving by Round')
        mergeHdrs = [['Result', 2], ['Score', 2]]
        headers = ['Round', 'Table', 'NS', 'EW', 'Board', 'Vul', 'Contract', 'By', 'Made', 'Down', 'NS', 'EW']
        sh = self.wb.create_sheet('By Round', 2)
        cStart = headers.index('Made')+1
        for h in mergeHdrs:
            sh.cell(1, cStart).value = h[0]
            sh.cell(1, cStart).font = self.noChangeFont
            sh.cell(1, cStart).alignment = self.centerAlign
            sh.merge_cells(f'{sh.cell(1,cStart).coordinate}:{sh.cell(1,cStart+h[1]-1).coordinate}')
            cStart += h[1]
        row = self.headerRow(sh, headers, 2)
        startRow = row
        for r in sorted(self.roundData.keys()): # round
            sh.cell(row, 1).value = r+1
            sh.cell(row, 1).alignment = self.centerAlign
            for t in sorted(self.roundData[r]): # table
                sh.cell(row, 2).value = t+1
                sh.cell(row, 3).value = self.pairN(self.roundData[r][t]['NS'])
                sh.cell(row, 4).value = self.pairN(self.roundData[r][t]['EW'])
                for b in self.roundData[r][t]['Board']:
                    sh.cell(row, 5).value = b+1
                    sh.cell(row, 6).value = f"{self.vulLookup(b)}"
                    for i in range(2,7):
                        sh.cell(row, i).alignment = self.centerAlign
                    row += 1
                for i in range(2,len(headers)+1):
                    sh.cell(row-1, i).border = self.bottomLine
            sh.cell(row-1,1).border = self.bottomLine
        return

    # Pickup Slip is PDF only
    # They are scores kept on "per round" basis.  There's a slip for each table for each round.
    # This is the ACBL style pickup slips with information pre-filled in.
    def Pickups(self):
        # Temporary data structure for ease of coding
        tables = {}
        for b,r in self.boardData.items():
            for v in r:
                if v[1] not in tables:
                    tables[v[1]] = {}
                if v[0] not in tables[v[1]]:
                    tables[v[1]][v[0]] = []
                tables[v[1]][v[0]].append({'NS': v[2], 'EW': v[3], 'Board': b})
        tblCols = []
        xMargin = self.pdf.margin
        hdrs = ['NS Score', 'Made', 'Down', 'NS Contract', 'By', 'Board', 'EW Contract', 'By', 'Made', 'Down', 'EW Socre']
        self.pdf.set_font(self.pdf.sansSerifFont, style='B', size=self.pdf.notePt)
        self.pdf.setHeaders(xMargin, hdrs, tblCols)
        bIdx = 0
        for t in sorted(tables.keys()):
            # the sit-out table
            for r in sorted(tables[t].keys()):
                nsPair = tables[t][r][0]['NS']
                ewPair = tables[t][r][0]['EW']
                if self.ifSitout(t, nsPair, ewPair):
                    continue
                if bIdx % 4 == 0:
                    self.pdf.add_page()
                    y = 2 * self.pdf.margin
                self.pdf.set_font(self.pdf.serifFont, style='B', size=self.pdf.headerPt)
                title = f"Pickup for Table {t+1}, Round {r+1}. NS: {self.pairN(nsPair)}, EW: {self.pairN(ewPair)}"
                self.printPickup(title, tables[t][r], tblCols, hdrs, xMargin, y)
                bIdx += 1
                y = self.pdf.sectionDivider(4, bIdx, xMargin)

        # People tend to lose pickup Slips.  They are given in the beginning for all rounds.
        # Print extra to fill a page or just 4 extras
        nExtra = 4 - bIdx % 4
        for _ in range(nExtra):
            if bIdx % 4 == 0:
                self.pdf.add_page()
                y = 2 * self.pdf.margin
            title = f"Table {" "*3}, Round {" "*3}, NS: {" "*3}, EW: {" "*3}"
            self.printPickup(title, ['']*len(tables[0][0]), tblCols, hdrs, xMargin, y)
            bIdx += 1
            y = self.pdf.sectionDivider(4, bIdx, xMargin)
        return

    def printPickup(self, title, boards, tblCols, hdrs, xMargin, y):
        y += self.pdf.lineHeight(self.pdf.font_size_pt)
        self.pdf.set_font(self.pdf.sansSerifFont, style='B', size=self.pdf.notePt)
        y = self.pdf.headerRow(xMargin, y, tblCols, hdrs, title)
        self.pdf.set_font(size=self.pdf.notePt)
        h = self.pdf.lineHeight(self.pdf.font_size_pt)
        y += h
        self.pdf.set_xy(xMargin, y)
        boardCol = hdrs.index('Board')
        for b in boards:
            for i in range(boardCol):
                self.pdf.cell(tblCols[i], h, text=f'', align='C', border=1)
            bText = f'{b["Board"]+1}' if type(b) != str else b
            self.pdf.cell(tblCols[5], h, text=bText, align='C', border=1)
            for i in range(boardCol+1,len(hdrs)):
                self.pdf.cell(tblCols[i], h, text=f'', align='C', border=1)
            y += h
            self.pdf.set_xy(xMargin, y)

    # Journal is for each pair to keep their own records
    # TD may collect them at the end to corroborate traveler or pickup slips
    def Journal(self):
        pairData = {}
        pairIdx = [2, 3]
        for b,r in self.boardData.items():  # board and rounds
            for v in r: # each round is (round, table, NS, and EW)
                for p in pairIdx: # NS and EW pairs in "v"
                    if v[p] not in pairData:
                        pairData[v[p]] = []
                    pairData[v[p]].append((b, v[0], v[1], v[2], v[3])) # (board, round, table, NS, EW)
        tblCols = []
        nPerPage = 2 if len(pairData[1]) < 18 else 1
        pIdx = 0
        xMargin = self.pdf.margin * 2
        hdrs = ['Board', 'Round', 'Sit-Out', 'EW', 'Contract', 'By', 'Made', 'Down', '8'*4, '8'*4]
        self.pdf.set_font(self.pdf.serifFont, style='B', size=self.pdf.headerPt)
        self.pdf.setHeaders(xMargin, hdrs, tblCols)
        hdrs[2] = 'NS'  # used "sit-out" to make sure sufficient width
        hdrs[8] = 'NS'
        hdrs[9] = 'EW'
        for p in sorted(pairData.keys()):
            if self.pairID(p) == self.SITOUT:
                continue
            if pIdx % nPerPage == 0:
                self.pdf.add_page()
                self.pdf.headerFooter()
                y = self.pdf.margin*2
            self.pdf.set_font(self.pdf.serifFont, style='B', size=self.pdf.headerPt)
            y = self.pdf.headerRow(xMargin, y, tblCols, hdrs ,f"Play Records for Pair {self.pairID(p)}")
            y += self.pdf.lineHeight(self.pdf.font_size_pt)
            self.pdf.set_font(size=self.pdf.linePt)
            h = self.pdf.lineHeight(self.pdf.font_size_pt)
            self.pdf.set_xy(xMargin, y)
            for v in sorted(pairData[p], key=lambda x: x[0]):
                self.pdf.cell(tblCols[0], h, text=f'{v[0]+1}', align='C', border=1)
                self.pdf.cell(tblCols[1], h, text=f'{v[1]+1}', align='C', border=1)
                self.pdf.cell(tblCols[2], h, text=f'{self.pairN(v[3])}', align='C', border=1)
                self.pdf.cell(tblCols[3], h, text=f'{self.pairN(v[4])}', align='C', border=1)
                for c in range(4,len(hdrs)):
                    self.pdf.cell(tblCols[c], h, text='', align='C', border=1)
                y += h
                self.pdf.set_xy(xMargin, y)
            pIdx += 1
            y = self.pdf.sectionDivider(nPerPage, pIdx, self.pdf.margin)
        return

    # Traveler goes with each board and "travel" among tables
    # Data {pair #: [(round, table, NS, EW), ...], ...}
    def Travelers(self):
        tblCols = []
        xMargin = 0.5
        hdrs = [self.SITOUT,'Contract', 'By', 'Made', 'Down', '8'*4, '8'*4, 'vs.']
        self.pdf.set_font(self.pdf.serifFont, style='B', size=self.pdf.headerPt)
        self.pdf.setHeaders(xMargin, hdrs, tblCols)
        hdrs[0] = 'NS'
        hdrs[5] = 'NS'
        hdrs[6] = 'EW'
        nPerPage = 4 if len(self.boardData[0]) <= 5 else 2 if len(self.boardData[0]) <= 12 else 1
        bIdx = 0
        for b,r in self.boardData.items():
            if bIdx % nPerPage == 0:
                self.pdf.add_page()
                y = 0.5
            y = self.printTraveler(xMargin, tblCols, hdrs, b, r, y)
            bIdx += 1
            if nPerPage > 1:
                y = self.pdf.sectionDivider(nPerPage, bIdx, xMargin)
        return

    # We may use this to print extra blank travelers.
    # Similar to Pickup slips
    def printTraveler(self, leftSide, tblCols, hdrs, bdNum, round, y):
        self.pdf.set_font(self.pdf.serifFont, style='B', size=self.pdf.headerPt)
        y = self.pdf.headerRow(leftSide, y, tblCols, hdrs, f'Travler for Board {bdNum+1}')
        y += self.pdf.lineHeight(self.pdf.font_size_pt)
        self.pdf.set_font(self.pdf.sansSerifFont, size=self.pdf.linePt)
        h = self.pdf.lineHeight(self.pdf.font_size_pt)
        for v in sorted(round, key=lambda x: x[2]):
            self.pdf.set_xy(leftSide, y)
            if type(v[2]) == str:
                self.pdf.cell(tblCols[0], h, text=v[2], align='C', border=1)
            else:
                self.pdf.cell(tblCols[0], h, text=f'{self.pairN(v[2])}', align='C', border=1)
            for c in range(1,len(hdrs)-1):
                    self.pdf.cell(tblCols[c], h, text='', align='C', border=1)
            self.pdf.cell(tblCols[len(hdrs)-1], h, text=f'{self.pairN(v[3])}', align='C', border=1)
            y += h
        return y

    # Print out the instructions for each table's movement card
    # (The table in the middle of the page)
    def Tables(self, nsTexts, ewTexts):
        tables = {}
        for b,r in self.boardData.items():
            for v in r:
                if v[1] not in tables:
                    tables[v[1]] = {}
                if v[0] not in tables[v[1]]:
                    tables[v[1]][v[0]] = []
                tables[v[1]][v[0]].append({'NS': v[2], 'EW': v[3], 'Board': b})
        hdrs = ['Round', 'NS', 'EW', 'Boards']
        tblCols = []
        xMargin = 0.5
        self.pdf.set_font(self.pdf.sansSerifFont, style='B', size=self.pdf.rosterPt)
        self.pdf.setHeaders(xMargin, hdrs, tblCols)
        bdsWidth = self.pdf.get_string_width('8'*self.decks*2+','*(self.decks-1)) + 0.25
        tblCols[3] = max(bdsWidth, tblCols[3])
        w = sum(tblCols)
        xMargin = (self.pdf.w - w) / 2
        tblHeight = (len(tables[0][0])+1) * self.pdf.pt2in(self.pdf.rosterPt)
        top = self.pdf.pt2in(self.pdf.bigPt) * 2.5 + 1
        compassTop = self.pdf.h - (self.pdf.pt2in(self.pdf.bigPt) * 5 + self.pdf.starRadius + 2)

        fontSize = self.pdf.rosterPt if len(tables) < 10 else self.pdf.bigPt
        for t in sorted(tables.keys()):
            if self.ifSitout(t, tables[t][0][0]['NS'], tables[t][0][0]['EW']):
                continue
            self.pdf.add_page()
            self.pdf.movementSheet()
            if compassTop > top + tblHeight:
                self.pdf.compass()
            self.pdf.tableAnchors(f"{t+1}")
            if nsTexts != None and ewTexts != None:
                self.pdf.inkEdgeText(nsTexts[t], ewTexts[t])
            self.pdf.set_font(self.pdf.sansSerifFont, style='B', size=fontSize)
            self.pdf.headerRow(xMargin, top, tblCols, hdrs, ' ')
            self.pdf.set_font(size=fontSize)
            y = self.pdf.get_y()
            h = self.pdf.lineHeight(self.pdf.font_size_pt);
            self.pdf.set_xy(xMargin, y + h)
            for r in sorted(tables[t].keys()):
                tRound = tables[t][r]
                self.pdf.cell(tblCols[0], h, text=f'{r+1}', align='C', border=1)
                self.pdf.cell(tblCols[1], h, text=f'{self.pairN(tRound[0]['NS'])}', align='C', border=1)
                self.pdf.cell(tblCols[2], h, text=f'{self.pairN(tRound[0]['EW'])}', align='C', border=1)
                bds = ""
                for b in tRound:
                    bds += f'{b['Board']+1},'
                self.pdf.cell(tblCols[3], h, text=bds[:-1], align='C', border=1)
                y += h
                self.pdf.set_xy(xMargin, y + h)

    # ID tag is for each person (not pair) to hold on to.
    # It gives extra information on next table and opponents.
    # It is in small font as the person can read it up-close.
    def idTags(self):
        idData = {}
        for rd, tbl in self.roundData.items(): # (round, table, NS, EW)
            for t, r in tbl.items():
                if r['NS'] not in idData:
                    idData[r['NS']] = []
                if r['EW'] not in idData:
                    idData[r['EW']] = []
                idData[r['NS']].append((rd, t, r['NS'], r['EW']))
                idData[r['EW']].append((rd, t, r['NS'], r['EW']))
        self.idTagsByData(idData)

    def idTagsByData(self, data):
        tags = 0
        colW = []
        hdrs = ['Round', 'Table', 'Seat', 'vs']
        self.pdf.setHeaders(0, hdrs, colW)
        # page can be portrait or landscape
        w = min(self.pdf.w,self.pdf.h)
        leftMargin = (w - 2 * sum(colW)) / 4
        cWidth = w / 2

        nTagsPage = 1 if len(data[1]) > 15 else (4 if len(data[1]) <= 8 else 2)
        for id in sorted(data.keys()):
            if self.pairID(id) == self.SITOUT:
                continue
            if tags % nTagsPage == 0:
                self.pdf.add_page(orientation='P') # no header/footer
                y = self.pdf.margin * 2
            rData = sorted(data[id], key=lambda x: x[0])    # by round
            for half in range(2):   # two identical tags for each person of the pair
                self.pdf.set_font(self.pdf.serifFont, style='B', size=self.pdf.headerPt)
                self.pdf.set_xy(leftMargin+cWidth*half, y)
                self.pdf.cell(text=f"Pair: {self.pairID(id)}")
                ty = y + self.pdf.lineHeight(self.pdf.font_size_pt)
                self.pdf.set_xy(leftMargin+cWidth*half, ty)
                self.pdf.set_font(self.pdf.sansSerifFont, style='B', size=self.pdf.smallPt)
                h = self.pdf.lineHeight(self.pdf.font_size_pt)
                for i in range(len(hdrs)):
                    self.pdf.cell(colW[i], h, text=hdrs[i], align='C', border=1)
                ty +=  h

                for r in rData:
                    opp,seat = (self.pairN(r[3]),'NS') if id == r[2] else (self.pairN(r[2]),'EW')
                    self.pdf.set_xy(leftMargin+cWidth*half, ty)
                    self.pdf.cell(colW[0], h, text=f"{r[0]+1}", align='C', border=1)
                    self.pdf.cell(colW[1], h, text=f"{r[1]+1}", align='C', border=1)
                    self.pdf.cell(colW[2], h, text=f"{seat}", align='C', border=1)
                    self.pdf.cell(colW[3], h, text=f"{opp}", align='C', border=1)
                    ty += h
            tags += 1
            y = self.pdf.sectionDivider(nTagsPage, tags, self.pdf.margin) + self.pdf.margin * 2
        return

    # Computing MPs
    # cIdx is the column on the left (one less) of the MP% columns
    # We have 4 MP "result" columns: 2 for % and 2 for Pts.
    # Then we have another 2 columns for "net" scores.
    # That's 6 columns, therefore the magic "7" is where the calculation area start
    #
    # "Net" is a simple formula so that no cell is blank, other than Averages
    # The calculation area is pair-wise comparisons to all opponents.  It's slightly esoteric.
    def computeMP(self, sh, cIdx, nPlayed, row, cursorRow, netIdx, calcStart=7):
        Win = 1.0
        Tie = 0.5
        Lost = 0.0
        # do twice, one for NS and another for EW
        for i in range(2):
            cStart = cIdx + calcStart + i*(nPlayed - 1)
            cEnd   = cStart + nPlayed - 2
            spread=f"{self.rc2a1(row, cStart)}:{self.rc2a1(row, cEnd)}"
            # The results are the computations from the calculation areas
            # "Non-comparisions" are blank cells and skipped by "COUNT"
            # Therefore the % is based on the times the board is actually played, not counting Averages
            # The Average cell is simply assigned as 50%
            sh.cell(row, cIdx+1+i).value = f"=IF(COUNT({spread})>0,{self.rc2a1(row, cIdx+3+i)}/COUNT({spread}),{Tie})"
            sh.cell(row, cIdx+1+i).number_format = sh.cell(row, cIdx+2).number_format = "0.00%"
            sh.cell(row, cIdx+3+i).value = f"=SUM({spread})"
            sh.cell(row, cIdx+3+i).number_format = "#0.00"

            # "cursorRow" is the counter within the current "play" group.  It goes from 0 to nPlayed - 1.
            # This comprehensive create the relative rows for *this* player to compare with.
            # By definition, there's only nPlayed - 1 opponents.
            opponents = [x - cursorRow for x in range(nPlayed) if x != cursorRow]
            n = nPlayed - 1
            # each column advanced (targetC) compares this row with one of the opponents (different rows)
            # So this is like 2-dimensional movements
            for rCmp in range(n):
                # if self is not a number, then blank out all comparisions
                # if the opponent is not a number, make that comparision blank
                # Otherwise, a win is 1 pt, tie 0.5, and lost 0.o
                cmpF = f"=IF(ISNUMBER({self.rc2a1(row, netIdx+i)}),IF(ISNUMBER({self.rc2a1(row+opponents[rCmp], netIdx+i)}),"
                cmpF += f"IF({self.rc2a1(row, netIdx+i)}>{self.rc2a1(row+opponents[rCmp], netIdx+i)},{Win},"
                cmpF += f'IF({self.rc2a1(row, netIdx+i)}={self.rc2a1(row+opponents[rCmp], netIdx+i)},{Tie},{Lost})),""),"")'
                targetC = cIdx+calcStart+rCmp+i*n
                sh.cell(row, targetC).value = cmpF

    # Net columns a convenient for MP/IMP computation.
    def computeNet(self, sh, row, raw, target):
        rawNS = self.rc2a1(row, raw)
        rawEW = self.rc2a1(row, raw+1)
        sh.cell(row, target).value = f'=IF(ISNUMBER({rawNS}),{rawNS},IF(ISNUMBER({rawEW}),-{rawEW},""))'
        sh.cell(row, target+1).value = f'=IF(ISNUMBER({rawEW}),{rawEW},IF(ISNUMBER({rawNS}),-{rawNS},""))'

    def computeIMP(self, sh, cIdx, nPlayed, row, cursorRow, netIdx, calcStart=9):
        calcStart += 2 * (nPlayed - 1)
        for i in range(2):
            cStart = cIdx + calcStart + i*(nPlayed - 1)
            cEnd   = cStart + nPlayed - 2
            spread=f"{self.rc2a1(row, cStart)}:{self.rc2a1(row, cEnd)}"
            sh.cell(row, cIdx+1+i).value = f'=AVERAGE({spread})'
            sh.cell(row, cIdx+1+i).number_format = sh.cell(row, cIdx+2).number_format = "#0.0"
            opponents = [x - cursorRow for x in range(nPlayed) if x != cursorRow]
            n = nPlayed - 1
            for rCmp in range(n):
                cmpF = f"=IF(AND(ISNUMBER({self.rc2a1(row, netIdx+i)}),ISNUMBER({self.rc2a1(row+opponents[rCmp], netIdx+i)})),"
                cmpF += f"VLOOKUP(ABS({self.rc2a1(row, netIdx+i)}-{self.rc2a1(row+opponents[rCmp], netIdx+i)}),'IMP Table'!$A$2:$C$26,3)"
                cmpF += f'*SIGN({self.rc2a1(row, netIdx+i)}-{self.rc2a1(row+opponents[rCmp], netIdx+i)}),0)'
                targetC = cIdx+calcStart+rCmp+i*n
                sh.cell(row, targetC).value = cmpF
        return

    # Common function to print a row of "headers" stylistically
    def boardSheetHeaders(self, sh, nTbl):
        # first row setup some spanning column headers
        mergeHdrs = [['Result', 2], ['Score', 2], ['IMP', 2], ['MP %', 2], ['MP Pts', 2], ['Net', 2],
               ['MP Calculation', nTbl*2 - 2],['IMP Calculation', nTbl*2 - 2]]

        headers = ['Board', 'Round', 'Table', 'NS', 'EW', 'Vul', 'Contract', 'By', 'Made', 'Down'] + ['NS', 'EW'] * 5
        cStart = headers.index('Made')+1
        for h in mergeHdrs:
            sh.cell(1, cStart).value = h[0]
            sh.cell(1, cStart).font = self.noChangeFont
            sh.cell(1, cStart).alignment = self.centerAlign
            sh.merge_cells(f'{sh.cell(1,cStart).coordinate}:{sh.cell(1,cStart+h[1]-1).coordinate}')
            cStart += h[1]
        headers += [['NS MP Scores', nTbl - 1], ['EW MP Scores', nTbl - 1], ['NS IMP Pair-wise', nTbl - 1], ['EW IMP Pair-wise', nTbl - 1]]
        row = self.headerRow(sh, headers, 2)
        return (row, headers)
        
    # Draw vertical lines at certain columns
    def boardVerticals(self, sh, headers, ntbl):
        vertical = [headers.index('Made')+5]
        vertical.append(vertical[-1] + 6)
        vertical.append(vertical[-1] + 2)
        vertical.append(vertical[-1] + (ntbl - 1)*2) 
        for c in vertical:
            for r in range(2,sh.max_row+1):
                bd = sh.cell(r, c).border
                sh.cell(r, c).border = Border(left=self.thinLine, bottom=bd.bottom)

    # Print tournament's meta information on a sheet
    def sheetMeta(self, sh, metaData):
        # Standard copyright notice
        sh.cell(1, 1).value = self.pdf.headerText
        sh.cell(1, 1).font = Font(size=10, italic=True, color="5DADE2")

        sh.cell(2, 1).value = metaData['Title']
        sh.cell(2, 1).font = self.HeaderFont
        sh.merge_cells(f'{sh.cell(2,1).coordinate}:{sh.cell(2,5).coordinate}')
        sh.cell(2, 1).alignment = self.centerAlign

        for row, info in enumerate(metaData['Info'], 3):
            sh.cell(row, 1).value = info[0]
            sh.cell(row, 1).font = self.HeaderFont
            sh.merge_cells(f'{sh.cell(row,1).coordinate}:{sh.cell(row,2).coordinate}')
            sh.cell(row, 3).value = info[1]
            sh.cell(row, 3).font = self.HeaderFont
        return row

    # Board tab references its data from the Round tab, for consistency
    # Then it compute the scores on the sheet
    def boardTab(self):
        self.log.debug('Saving by Board')
        sh = self.wb.create_sheet('By Board', 1)
        row, headers = self.boardSheetHeaders(sh, self.tables)
        rGap = self.tables * self.decks    # Number of rows between each round
        for b in sorted(self.boardData.keys()):
            sh.cell(row, 1).value = b+1     # board #
            sh.cell(row, 1).alignment = self.centerAlign
            nPlayed = len(self.boardData[b])    # # of times this board was played
            cursorRow = 0
            for r in sorted(self.boardData[b], key=lambda x: x[2]): # (round, table, NS, EW)
                sh.cell(row, 2).value = f"='By Round'!{self.rc2a1(r[0] * rGap + 3, 1)}"
                tBase = r[0] * rGap + r[1] * self.decks + 3
                sh.cell(row, 3).value = f"='By Round'!{self.rc2a1(tBase, 2)}"
                sh.cell(row, 4).value = f"='By Round'!{self.rc2a1(tBase, 3)}"
                sh.cell(row, 5).value = f"='By Round'!{self.rc2a1(tBase, 4)}"
                tBase += b % self.decks
                for i in range(6, len(headers)+1):
                    c = f"'By Round'!{self.rc2a1(tBase, i)}"
                    sh.cell(row, i).value = f'=IF(ISBLANK({c}),"",{c})'
                for i in range(2,7):
                    sh.cell(row, i).alignment = self.centerAlign

                cIdx = headers.index('Made')+4
                nIdx = cIdx + 7
                self.computeNet(sh, row, cIdx-1, nIdx)
                self.computeIMP(sh, cIdx, nPlayed, row, cursorRow, nIdx)
                self.computeMP(sh, cIdx+2, nPlayed, row, cursorRow, nIdx)
                if self.fake:
                    self.fakeScore(sh, row, cIdx-1)
                row += 1
                cursorRow += 1
            for c in range(len(headers)+(self.tables-1)*4-4):
                sh.cell(row-1, c+1).border = self.bottomLine
        self.boardVerticals(sh, headers, self.tables)
        return
    
    # Some simple validity checks
    # boardData and roundData are generated from one, or the other.
    # Not checking if they are consistent to each other.
    def checkBoardData(self):
        if not hasattr(self, 'boardData') or len(self.boardData) <= 0:
            return False
        pData = {}
        # No one can play the same board more than once
        for b,t in self.boardData.items():
            for tbl in t:
                # (round table ns ew)
                for s in [2, 3]:
                    if tbl[s] not in pData:
                        pData[tbl[s]] = {'Boards': [], 'Against': []}
                    if b not in pData[tbl[s]]['Boards']:
                        pData[tbl[s]]['Boards'].append(b)
                    else:
                        raise ValueError('Same Baord', tbl[s], b+1)

        # No pair can meet anothe pair more than once
        sides = ['NS', 'EW']
        for t in self.roundData.values():
            for tbl in t.values():
                for s in [0,1]:
                    if tbl[sides[s]] not in pData[tbl[sides[1-s]]]['Against']:
                        pData[tbl[sides[1-s]]]['Against'].append(tbl[sides[s]])
                    else:
                        raise ValueError('Same Pair', tbl[sides[1-s]], tbl[sides[s]])

        # There are some "soft" rules
        # Players should play the same boards, at least the same number of boards
        # Boards should not appear on different tables for the same round
        # These checks must allow exceptions.  Harder to code.  Not sure worth the efforts.
        return True