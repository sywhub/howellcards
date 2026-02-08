#!/usr/bin/env python3
# Generate Mitchell movements
#   A PDF with Roster and Score sheets
#   An Excel spreadsheet to enter the results and calculate the scores
#
# A 4-table Mitchell use "Square" arrangement, found at MIT web site
#
import argparse
import logging
from maininit import setlog
from openpyxl import Workbook
from openpyxl.styles import Font
import pdf
from docset import PairGames
import datetime

# Pairs are internally numbered 1,3,5,... for EW pairs and 2,4,6,... for NS
# Pair 0 is the sit-out phantom pair
# Externally, they are number 1 to n for both NS and EW sides
class Mitchell(PairGames):
    def __init__(self, log, p, b, f):
        super().__init__(log)
        self.pairs = p
        self.decks = b
        self.tables = (self.pairs + 1) // 2
        self.oddPairs = self.pairs % 2 == 1
        self.fake = f
        self.pdf = pdf.PDF()
        self.wb = Workbook()

        self.pdf.HeaderFooterText(f'{self.notice} {datetime.date.today().strftime("%b %d, %Y")}.',
            f'Mitchell Tournament: {(self.pairs+1)//2} Tables, {self.decks} Boards per round')

    # identify whether the pair is NS or EW
    def pairSide(self, n):
        return ['NS', 'EW'][n % 2]

    # translate internal pair number to external
    def pairN(self, n):
        if n == 0:
            return self.SITOUT
        return n // 2 + (0 if n % 2 == 0 else + 1)

    # identify the side of the pair
    def pairID(self, n):
        return f"{self.pairSide(n)} {self.pairN(n)}" if n != 0 else self.SITOUT

    # assign NS pair number
    def NSPair(self, r, t):
        n = (t+1) * 2
        return n if n <= self.pairs else 0

    # assign EW pair number
    def EWPair(self, r, t):
        x = (self.tables - r) % self.tables + t
        x %= self.tables
        return x * 2 + 1
    
    def boardIdx(self, r, t):
        return ((r + t) % self.tables) * self.decks

    def ifSitout(self, t, ns, ew):
        return ns == 0

    def main(self):
        self.log.debug('Main goes')
        # initData must be the first one
        self.initData()
        self.pdf.instructions(self.log, 'mitchellInstructions.txt')
        self.roster()
        self.results()
        self.roundTab()
        self.boardTab()
        self.IMPTable() # static sheet
        self.ScoreTable()   # static sheet, produced to aid human TD, not used elsewhere.
        self.idTags()  # PDF only
        self.setTableTexts()  # PDF only
        self.Pickups()  # PDF only
        self.Travelers()  # PDF only
        self.Journal()  # PDF only
        self.save()
        return


    # Generate "boardData" and "roundData"
    def initData(self):
        self.boardData = {}
        if self.pairs == 8:
            self.loadSquare()   # squaure Mitchell
        elif self.tables % 2 == 0: 
            self.loadEven() # self.pairs in [11, 12, 15, 16]
        else:  # standard Mitchell
            for r in range(self.tables): # round
                for t in range(self.tables): # table
                    b = self.boardIdx(r, t)
                    for bset in range(self.decks):
                        if (b + bset) not in self.boardData:
                            self.boardData[b+bset] = []
                        self.boardData[b+bset].append([r, t, self.NSPair(r, t), self.EWPair(r, t)])
        self.initRounds()

    def roster(self):
        self.log.debug('Roster sheet and PDF')
        self.rosterSheet()
        self.rosterPDF()
    
    # a notice on public domain
    # Then the meta info about this tournament
    # Last a list of names for pairs
    def rosterSheet(self):
        ws = self.wb.active # the first tab
        ws.title = 'Roster'

        tourney = {'Title':"Mitchell Tournament Scoring",
            'Info': [['Pairs', self.pairs], ['Tables', self.tables], ['Boards per Round', self.decks]]}

        row = self.sheetMeta(ws, tourney) + 2
        toN = self.pairs + (1 if self.oddPairs else 0)
        for s in range(2):
            ws.cell(row, 1).value =  f'{['NS', 'EW'][s]} Pairs'
            ws.cell(row, 1).font = self.HeaderFont
            ws.cell(row, 1).alignment = self.centerAlign
            ws.merge_cells(f'{ws.cell(row,1).coordinate}:{ws.cell(row,3).coordinate}')
            ws.cell(row, 4).value = 'MP'
            ws.cell(row, 4).font = self.HeaderFont
            ws.cell(row, 4).alignment = self.centerAlign
            ws.cell(row, 5).value = 'IMP'
            ws.cell(row, 5).font = self.HeaderFont
            ws.cell(row, 5).alignment = self.centerAlign
            row += 1
            avgStart = row  # remember this row
            for p in range(s, toN, 2):
                pName = self.pairN(p+1)
                if pName == self.SITOUT:
                    continue
                ws.cell(row, 1).font = self.HeaderFont
                ws.cell(row, 1).alignment = self.centerAlign
                ws.cell(row, 1).value = pName
                ws.cell(row, 2).value = self.placeHolderName()
                ws.cell(row, 3).value = self.placeHolderName()
                row += 1

            # draw a line
            for i in range(6):
                ws.cell(row-1, i+1).border = self.bottomLine

            ws.cell(row, 3).value = 'Average'
            ws.cell(row,3).font = self.noChangeFont

            ws.cell(row, 4).value = f'=AVERAGE({self.rc2a1(avgStart, 4)}:{self.rc2a1(row-1,4)})'
            ws.cell(row, 5).value = f'=SUM({self.rc2a1(avgStart, 6)}:{self.rc2a1(row-1,6)})'
            ws.cell(row,4).number_format = "0.00%"
            ws.cell(row,5).number_format = "#0.0"
            ws.cell(row,4).font = self.noChangeFont
            ws.cell(row,5).font = self.noChangeFont
            row += 2
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 30
        
    def rosterPDF(self):
        y = self.meta()
        self.pdf.set_font(self.pdf.serifFont, style='B', size=self.pdf.rosterPt) 
        h = self.pdf.lineHeight(self.pdf.font_size_pt)
        title = 'Player Pairs'
        x = self.pdf.setHCenter(self.pdf.get_string_width(title))
        y += 2 * h
        self.pdf.set_xy(x, y)
        self.pdf.cell(text=title)
        widths = [1, 2, 2]
        y +=  h
        leftM = (self.pdf.w - sum(widths)) / 2
        self.pdf.set_xy(leftM, y)
        self.pdf.set_font(self.pdf.sansSerifFont, size=(self.pdf.bigPt if self.pairs < 19 else self.pdf.linePt)) 
        h = self.pdf.lineHeight(self.pdf.font_size_pt)
        start = 2
        for s in ['NS', 'EW']:
            self.pdf.set_font(style='BI')
            self.pdf.cell(5, h, text=f'{s} Pairs', align='L')
            self.pdf.set_font(style='')
            y += h
            self.pdf.set_xy(leftM, y)
            for p in range(start, self.pairs + 1, 2):
                self.pdf.cell(widths[0], h, text=f'{self.pairN(p)}', align='C', border=1)
                self.pdf.cell(widths[1], h, text='', align='C', border=1)
                self.pdf.cell(widths[2], h, text='', align='C', border=1)
                y += h
                self.pdf.set_xy(leftM, y)
            start -= 1
        return

    # roster shows meta info first
    # Not doing meta sheet
    def meta(self):
        self.log.debug('Meta')
        meta = {'Title': 'Mitchell Tournament', 'Info': []}
        meta['Info'].append(('Pairs', self.pairs))
        meta['Info'].append(('Tables', self.tables))
        if self.tables % 2 == 0:
            meta['Info'].append(('Relay between ', f"{self.tables // 2} and {self.tables // 2 + 1}"))
        meta['Info'].append(('Rounds', self.pairs // 2 - 1))
        meta['Info'].append(('Boards per round', self.decks))
        self.pdf.add_page()
        self.pdf.headerFooter()
        return self.pdf.meta(meta)

    def setTableTexts(self):
        self.log.debug('Setting Table borders')
        if self.pairs == 8:
            # The trade-off of square is the iregularity of movements
            ewText = ['R2 to T2/EW, R3 to T3/EW, R4 to T2/EW',
                        'R2 to T1/EW, R3 to T4/EW, R4 to T1/EW',
                        'R2 to T4/EW, R3 to T1/EW, R4 to T4/EW',
                        'R2 to T3/EW, R3 to T2/EW, R4 to T3/EW']
            nsText = ['Stay here. Boards: R2 to T4, R3 to T2, R4 to T4',
                        'Stay here. Boards: R2 to T3, R3 to T1, R4 to T3', 
                        'Stay here. Boards: R2 to T2, R3 to T4, R4 to T2', 
                        'Stay here. Boards: R2 to T1, R3 to T3, R4 to T1'] 
        else:
            nsText = []
            ewText = []
            for t in range(self.tables):
                ewText.append(f'Move to Table {t+2 if t < 3 else 1} EW')
                if self.tables % 2 == 0 and t == self.tables // 2:
                    nsText.append(f'Stay Here, Boards to Relay')
                else:
                    nsText.append(f'Stay Here, Boards to T{t if t > 0 else 4}')
        self.Tables(nsText, ewText)

    # Square arrangement is not programatic.
    def loadSquare(self):
        self.log.debug('Load Square data')
        self.sqSetup = {
            # Primary key is the table number
            # Pair numbering in this data is separated by NS/EW
            # Each "board set" is n boards, as dedicated by command line argument
            0: [{'Round': 0, 'NS': 1, 'EW': 1, 'Board': 0},   # round, NS, EW, boardSet #
                {'Round': 1, 'NS': 1, 'EW': 2, 'Board': 3},
                {'Round': 2, 'NS': 1, 'EW': 4, 'Board': 2},
                {'Round': 3, 'NS': 1, 'EW': 3, 'Board': 1},],
            1: [{'Round': 0, 'NS': 2, 'EW': 2, 'Board': 1},
                {'Round': 1, 'NS': 2, 'EW': 1, 'Board': 2},
                {'Round': 2, 'NS': 2, 'EW': 3, 'Board': 3},
                {'Round': 3, 'NS': 2, 'EW': 4, 'Board': 0},],
            2: [{'Round': 0, 'NS': 3, 'EW': 3, 'Board': 2},
                {'Round': 1, 'NS': 3, 'EW': 4, 'Board': 1},
                {'Round': 2, 'NS': 3, 'EW': 2, 'Board': 0},
                {'Round': 3, 'NS': 3, 'EW': 1, 'Board': 3},],
            3: [{'Round': 0, 'NS': 4, 'EW': 4, 'Board': 3},
                {'Round': 1, 'NS': 4, 'EW': 3, 'Board': 0},
                {'Round': 2, 'NS': 4, 'EW': 1, 'Board': 1},
                {'Round': 3, 'NS': 4, 'EW': 2, 'Board': 2}]}
        self.boardData = {}
        for t,tbl in self.sqSetup.items():
            for r in tbl:
                r['Board'] = [r['Board']*self.decks + x for x in range(self.decks)]
                r['NS'] = r['NS'] * 2
                r['EW'] = (r['EW'] - 1) * 2 + 1
                for b in r['Board']:
                    if b not in self.boardData:
                        self.boardData[b] = []
                    self.boardData[b].append([r['Round'], t, r['NS'], r['EW']])

    def loadEven(self):
        self.roundData = {}
        for r in range(self.tables):
            self.roundData[r] = {}
            for t in range(self.tables):
                self.roundData[r][t] = []
                bIdx = t if t <= 2 else t + 1
                bIdx += r
                if bIdx > self.tables:
                    bIdx -= self.tables + 1
                blist = [self.decks*bIdx+x for x in range(self.decks)]
                self.roundData[r][t] = {'NS': self.NSPair(r, t), 'EW': self.EWPair(r,t), 'Board': blist}
        self.boardData = {}
        for r,tbl in self.roundData.items():
            for t,d in tbl.items():
                for b in d['Board']:
                    if b not in self.boardData:
                        self.boardData[b] = []
                    self.boardData[b].append([r, t, d['NS'], d['EW']])


    def results(self):
        self.log.debug('Add results to Roster')
        sh = self.wb['Roster']
        nRows = 0
        row = 8
        divident = len(self.roundData) * len(self.roundData[0][0]['Board'])

        for b in self.boardData.values():
            nRows += len(b)
        for s in range(2):
            toN = self.pairs + (1 if self.oddPairs else 0)
            for p in range(s, toN, 2):
                pName = self.pairN(p+1)
                if pName == self.SITOUT:
                    continue
                ifRange = f"'By Board'!{self.rc2a1(3, 4+s)}:{self.rc2a1(3+nRows,4+s)}"
                impRange = f"'By Board'!{self.rc2a1(3, 12+s)}:{self.rc2a1(3+nRows,12+s)}"
                sumRange = f"'By Board'!{self.rc2a1(3, 14+s)}:{self.rc2a1(3+nRows,14+s)}"
                sh.cell(row,4).value=f"=SUMIF({ifRange},\"=\"&{self.rc2a1(row, 1)},{sumRange})/{divident}"
                sh.cell(row,5).value=f"=SUMIF({ifRange},\"=\"&{self.rc2a1(row, 1)},{impRange})"
                sh.cell(row,4).number_format = "0.00%"
                sh.cell(row,5).number_format = "#0.0"
                row += 1
            row += 3

    # Output into filesystem
    def save(self):
        import os
        here = os.path.dirname(os.path.abspath(__file__))
        fn = f'{here}/../mitchell{self.pairs}x{self.decks}{"xF" if self.fake else ""}'
        self.log.debug(f'Save files: {fn}')
        self.wb.save(f'{fn}.xlsx')
        self.pdf.output(f'{fn}.pdf')
        print(f'Saved {fn}.{{xlsx,pdf}}')


if __name__ == '__main__':
    log = setlog('mitchell', None)
    def mitchell_check(value):
        ivalue = int(value)
        if ivalue in [11,15,16]:
            raise argparse.ArgumentTypeError(f"Cannot have even number of tables")
        return ivalue

    parser = argparse.ArgumentParser()
    parser.add_argument('-d', '--debug', type=str, default='INFO', help='Debug level, INFO, DEBUG, ERROR')
    parser.add_argument('-b', '--boards', type=int, choices=range(1,7), default=4, help='Boards per round')
    parser.add_argument('-p', '--pair', type=int, choices=range(8,25), default=8, help='Number of pairs')
    parser.add_argument('-f', '--fake', type=bool, default=False, help='Fake scores to test the spreadsheet')
    args = parser.parse_args()
    for l in [['INFO', logging.INFO], ['DEBUG', logging.DEBUG], ['ERROR', logging.ERROR]]:
        if args.debug.upper() == l[0]:
            log.setLevel(l[1])
            break
    mitchell = Mitchell(log, args.pair, args.boards, args.fake)
    mitchell.main()
