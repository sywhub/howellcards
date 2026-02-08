#!/usr/bin/env python3
# Generate a team match setup
#   A PDF with Roster and Score sheets
#   An Excel spreadsheet to enter the results and calculate the scores
#
# 4 Pairs (1, 2, 3, 4) are arranged into 2 teams (1 and 2, 3 and 4) and play each other.
#   They sit at 2 tables: {NS: Team 1, EW: 3} {NS: Team 4, EW: Team 2}.
#       Each table play n boards, then they exchange boards to play again.
#       Then they exchage the oppoenents:{NS: Team 1, EW: 4} {NS: Team 3, EW: 2}.
#       Play as before again.
#       Technically, that's 4 rounds.
#       The scoring diffreences are converted into IMPs and the one with higher IMP wins.
#
#   Now, swap team composition (1 and 3, 2 and 4).  Repeat as above.  That's anothe 4 rounds.
#   Swap again (1 and 4, 2 and 3).  Again.
#
#   That's total of 12 rounds.  But really 3 team matches.
#
import argparse
import logging
import pdf
import datetime
from openpyxl import Workbook
from docset import PairGames
from maininit import setlog

class TeamMatch(PairGames):
    def __init__(self, log):
        super().__init__(log)
        self.pdf = pdf.PDF()
        self.wb = Workbook()
        self.decks = 4
        self.pairs = 4
        self.pdf.HeaderFooterText(f'{self.notice} {datetime.date.today().strftime("%b %d, %Y")}.',
           f'Team Match {self.decks} per Round')

    # record metadata
    def setup(self, boards, fake):
        self.decks = boards
        self.fake = fake
        self.initData()

    def pairN(self, n):
        return n

    def pairID(self, n):
        return f"{n}"

    # Setup boardData and roundData for parent class methods
    def initData(self):
        self.roundData = {
            0: {0: {'NS': 1, 'EW': 3, 'Board': 0}, 1: {'NS': 4, 'EW': 2, 'Board': 1}},
            1: {0: {'NS': 1, 'EW': 3, 'Board': 1}, 1: {'NS': 4, 'EW': 2, 'Board': 0}},
            2: {0: {'NS': 1, 'EW': 4, 'Board': 2}, 1: {'NS': 3, 'EW': 2, 'Board': 3}},
            3: {0: {'NS': 1, 'EW': 4, 'Board': 3}, 1: {'NS': 3, 'EW': 2, 'Board': 2}},

            4: {0: {'NS': 1, 'EW': 4, 'Board': 4}, 1: {'NS': 2, 'EW': 3, 'Board': 5}},
            5: {0: {'NS': 1, 'EW': 4, 'Board': 5}, 1: {'NS': 2, 'EW': 3, 'Board': 4}},
            6: {0: {'NS': 1, 'EW': 2, 'Board': 6}, 1: {'NS': 4, 'EW': 3, 'Board': 7}},
            7: {0: {'NS': 1, 'EW': 2, 'Board': 7}, 1: {'NS': 4, 'EW': 3, 'Board': 6}},

            8: {0: {'NS': 1, 'EW': 2, 'Board': 8}, 1: {'NS': 3, 'EW': 4, 'Board': 9}},
            9: {0: {'NS': 1, 'EW': 2, 'Board': 9}, 1: {'NS': 3, 'EW': 4, 'Board': 8}},
            10: {0: {'NS': 1, 'EW': 3, 'Board': 10}, 1: {'NS': 2, 'EW': 4, 'Board': 11}},
            11: {0: {'NS': 1, 'EW': 3, 'Board': 11}, 1: {'NS': 2, 'EW': 4, 'Board': 10}}}

        for r in self.roundData.keys():
            for t in self.roundData[r].keys():
                self.roundData[r][t]['Board'] = self.boardList(self.roundData[r][t]['Board'])
        for r,t in self.roundData.items():
            for tbl, p in t.items():
                for b in p['Board']:
                    if b not in self.boardData:
                        self.boardData[b] = []
                    self.boardData[b].append([r, tbl, p['NS'], p['EW']])
        

    # Roster sheet
    # The roster tab also shows the tournament results
    def rosterSheet(self):
        ws = self.wb.active # the first tab
        ws.title = 'Roster'
        metaData = {'Title': "Team Match",
                    'Info': [["Pairs", self.pairs], ["Rounds", len(self.roundData)], ["Boards per Round", self.decks]]}
        row = self.sheetMeta(ws, metaData)
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 30
        row += 2
        ws.cell(row, 1).value =  'Pairs'
        ws.cell(row, 1).font = self.HeaderFont
        ws.cell(row, 1).alignment = self.centerAlign
        ws.merge_cells(f'{ws.cell(row,1).coordinate}:{ws.cell(row,3).coordinate}')
        ws.cell(row, 4).value = 'IMP'
        ws.cell(row, 4).font = self.HeaderFont
        ws.cell(row, 4).alignment = self.centerAlign
        row += 1
        nBoards = len(self.boardData)
        for pair in range(4):
            ws.cell(row, 1).font = self.HeaderFont
            ws.cell(row, 1).alignment = self.centerAlign
            ws.cell(row, 1).value = pair+1
            ws.cell(row, 2).value = self.placeHolderName()
            ws.cell(row, 3).value = self.placeHolderName()
            sum = f"=SUMIF('By Board'!{self.rc2a1(3,4)}:{self.rc2a1(3+nBoards,4)},\"=\"&{self.rc2a1(row,1)},'By Board'!{self.rc2a1(3,12)}:{self.rc2a1(3+nBoards,12)})"
            sum += f"+SUMIF('By Board'!{self.rc2a1(3,5)}:{self.rc2a1(3+nBoards,5)},\"=\"&{self.rc2a1(row,1)},'By Board'!{self.rc2a1(3,13)}:{self.rc2a1(3+nBoards,13)})"
            ws.cell(row, 4).value = sum
            row += 1
        for c in range(4):
            ws.cell(row-1,c+1).border = self.bottomLine
        ws.cell(row,3).value='Sum'
        ws.cell(row,4).value=f'=SUM({self.rc2a1(3,4)}:{self.rc2a1(7,4)})'
        ws.cell(row,3).font = self.noChangeFont
        ws.cell(row,4).font = self.noChangeFont
        
    # simple sign-up sheet, PDF
    def rosterPDF(self):
        self.pdf.add_page()
        self.pdf.headerFooter()
        self.pdf.set_font(style='BI', size=self.pdf.rosterPt, family=self.pdf.serifFont)
        self.pdf.set_y(self.pdf.margin + self.pdf.lineHeight(self.pdf.font_size_pt) * 2)
        self.pdf.cell(w=self.pdf.epw, text='Pair Signup', align='C')
        self.pdf.set_y(self.pdf.get_y() + self.pdf.lineHeight(self.pdf.font_size_pt) * 2)
        
        pw = self.pdf.get_string_width('Pair'+'8'*2) + 0.25
        nameW = (self.pdf.w - pw - 4 * self.pdf.margin) / 2
        self.pdf.set_font(style='', family=self.pdf.sansSerifFont)
        ht = self.pdf.lineHeight(self.pdf.font_size_pt)

        for p in range(4):
            self.pdf.set_x(self.pdf.margin*2)
            self.pdf.cell(w=pw, h=ht, text=f'Pair {p+1}', align='C', border=1)
            self.pdf.cell(w=nameW, h=ht, text='', border=1)
            self.pdf.cell(w=nameW, h=ht, text='', border=1)
            self.pdf.ln()

    def boardSheetHeaders(self, sh, nTbl):
        # first row setup some spanning column headers
        mergeHdrs = [['Score', 2], ['IMP', 2], ['Net', 2]]

        headers = ['Board', 'Round', 'Table', 'NS', 'EW', 'Vul', 'Contract', 'By', 'Result'] + ['NS', 'EW'] * 3
        cStart = headers.index('Result') + 2
        for h in mergeHdrs:
            sh.cell(1, cStart).value = h[0]
            sh.cell(1, cStart).font = self.noChangeFont
            sh.cell(1, cStart).alignment = self.centerAlign
            sh.merge_cells(f'{sh.cell(1,cStart).coordinate}:{sh.cell(1,cStart+h[1]-1).coordinate}')
            cStart += h[1]
        row = self.headerRow(sh, headers, 2)
        return (row, headers)
        
    # Table of boards played, no PDF equivalent
    # Team matches are always IMP and 2 tables.  It always uses traveler.
    def Boards(self):
        self.log.debug('Saving by Board')
        sh = self.wb.create_sheet('By Board', 1)
        row, headers = self.boardSheetHeaders(sh, 2)
        for b in sorted(self.boardData.keys()):
            sh.cell(row, 1).value = b+1     # board #
            sh.cell(row, 1).alignment = self.centerAlign
            cursorRow = 0
            for r in self.boardData[b]: # (round, table, NS, EW)
                sh.cell(row, 2).value = r[0]+1  # round
                sh.cell(row, 3).value = r[1]+1  # table
                sh.cell(row, 4).value = r[2]    # NS
                sh.cell(row, 5).value = r[3]    # EW
                sh.cell(row, 6).value = self.vulLookup(b)
                for i in range(2,7):
                    sh.cell(row, i).alignment = self.centerAlign

                cIdx = headers.index('Result')+3
                nIdx = cIdx + 3
                self.computeNet(sh, row, cIdx-1, nIdx)
                self.computeIMP(sh, cIdx, 2, row, cursorRow, nIdx, -1)  # put *here*
                row += 1
                cursorRow += 1
            if self.fake:
                self.fakeScore(sh, row-2, cIdx-1, 1.0)
                self.fakeScore(sh, row-1, cIdx-1, 1.0)
            for c in range(nIdx+1):
                sh.cell(row-1,c+1).border = self.bottomLine
        
    # Output into filesystem
    def save(self):
        import os
        here = os.path.dirname(os.path.abspath(__file__))
        fn = f'{here}/../teammatchx{self.decks*2}'
        self.wb.save(f'{fn}.xlsx')
        self.pdf.output(f'{fn}.pdf')
        print(f'Saved {fn}.{{xlsx,pdf}}')

    # Orchestrator
    def match(self):
        self.pdf.instructions(self.log, "teaminstructions.txt")
        self.rosterSheet()
        self.rosterPDF()
        self.Boards()
        self.IMPTable()
        self.ScoreTable()
        self.idTags()
        self.Travelers()  # PDF only
        self.Journal()  # pdf only
        self.save()
        return

if __name__ == '__main__':
    log = setlog('team', None)

    parser = argparse.ArgumentParser()
    parser.add_argument('-d', '--debug', type=str, default='INFO', help='Debug level, INFO, DEBUG, ERROR')
    parser.add_argument('-b', '--boards', type=int, choices=range(1,5), default=2, help='Number of boards per round')
    parser.add_argument('-f', '--fake', type=bool, default=False, help='Fake scores to test the spreadsheet')
    args = parser.parse_args()
    for l in [['INFO', logging.INFO], ['DEBUG', logging.DEBUG], ['ERROR', logging.ERROR]]:
        if args.debug.upper() == l[0]:
            log.setLevel(l[1])
            break
    team = TeamMatch(log)
    # A match has n rounds, each round has m boards, divided into two halves, each half of the boards
    team.setup(boards=args.boards, fake=args.fake)
    team.match()